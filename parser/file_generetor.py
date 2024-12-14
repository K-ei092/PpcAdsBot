from openpyxl import Workbook
from bs4 import BeautifulSoup
import logging
from urllib.parse import urlparse


logger = logging.getLogger(__name__)


class FileGenerator:
    def __init__(self, advertising_seo: str, input_region: str, name_file: str):
        self.advertising_seo = advertising_seo
        self.input_request = ''
        self.input_region = input_region
        self.name_file = name_file
        self.count_row = 2
        self.wb = Workbook()
        self.ws = self.wb.active
        self.setup_workbook()
        self.site_weight = {
            'СП-1-1': 160, 'СП-1-2': 100, 'СП-1-3': 75, 'СП-1-4': 65, 'СП-1-5': 55,
            'М-1-1': 15, 'М-1-2': 5, 'М-1-3': 3, 'М-1-4': 2, 'М-1-5': 2, 'М-1-6': 2,
            'СП-2-1': 1, 'СП-2-2': 1, 'СП-2-3': 1, 'СП-2-4': 1, 'СП-2-5': 1,
            'М-2-1': 0.5, 'М-2-2': 0.5, 'М-2-3': 0.5, 'М-2-4': 0.5, 'М-2-5': 0.5, 'М-2-6': 0.5
                   }

    def setup_workbook(self):
        """Настраивает заголовки и ширину столбцов в рабочей книге."""
        headers = ['порядок выдачи', 'поисковый запрос', 'вес сайта', 'вид', 'адрес сайта', 'заголовок', 'описание']
        self.ws.append(headers)
        column_widths = [15, 20, 10, 10, 25, 30, 30, 20, 15, 20, 15, 20, 15, 20, 15]
        for index, width in enumerate(column_widths, start=1):
            self.ws.column_dimensions[self.ws.cell(row=1, column=index).column_letter].width = width
        self.wb.save(self.name_file)
        logger.info("Настроены заголовки и ширина столбцов в рабочей книге")

    def generate_file(self, response: str, input_request: str) -> str | None:
        """Основной метод для генерации файла."""
        self.input_request = input_request
        self.count_row = self.process_response(response, self.count_row)
        logger.info("Данные записаны в файл")
        self.wb.save(self.name_file)
        return self.name_file

    def process_response(self, response: str, count_row: int) -> int:
        """Обрабатывает ответ и заполняет рабочую книгу."""
        soup = BeautifulSoup(response, 'xml')
        logger.info("Приготовлен BS4")
        if soup:
            return self.parse_ads(soup, count_row)

    def parse_ads(self, soup, count_row: int) -> int:
        """Парсит рекламу и SEO из ответа и записывает в рабочую книгу."""
        search = self.advertising_seo
        topads = soup.find('topads')
        bottomads = soup.find('bottomads')
        seo = soup.find('grouping')
        logger.info("Собрана и реклама и SEO")

        namber_page = str(soup.find('page'))[-8]

        if search in {'Реклама', 'Реклама + SEO'}:
            count_row = self.parse_ads_section(topads, count_row, namber_page, 'СП') + 1
            count_row = self.parse_ads_section(bottomads, count_row, namber_page, 'М') + 1
            logger.info("Реклама записана в рабочую книгу")

        if search in {'SEO', 'Реклама + SEO'}:
            count_row = self.parse_seo_section(seo, count_row, namber_page) + 1
            logger.info("SEO записано в рабочую книгу")

        return count_row

    def parse_ads_section(self, ads, count_row: int, namber_page: str, prefix: str) -> int:
        """Парсит раздел рекламы и записывает данные в рабочую книгу."""
        if ads:
            for index, ads in enumerate(ads.find_all('query'), start=1):
                try:
                    self.ws.cell(row=count_row, column=1, value=f'{prefix}-{namber_page}-{index}')
                    self.ws.cell(row=count_row, column=2, value=self.input_request)
                    self.ws.cell(row=count_row, column=3, value=self.site_weight[f'{prefix}-{namber_page}-{index}'])
                    self.ws.cell(row=count_row, column=4, value='Реклама')
                    self.ws.cell(row=count_row, column=5, value=str(ads.find('url'))[5:-6])
                    self.ws.cell(row=count_row, column=6, value=self._check_tag(str(ads.find('title'))[7:-8]))
                    self.ws.cell(row=count_row, column=7, value=self._check_tag(str(ads.find('snippet'))[9:-10]))
                    self.parse_sitelinks(ads, count_row)
                    count_row += 1
                    logger.info(f"записана реклама - {ads}, по индексу - {prefix}-{namber_page}-{index}")
                except Exception as ex:
                    logger.exception(ex)
                    self.handle_parse_error(count_row, prefix, namber_page)
                    count_row += 1
        return count_row

    def parse_seo_section(self, seo, count_row: int, namber_page: str) -> int:
        """Парсит раздел SEO и записывает данные в рабочую книгу."""
        if seo:
            for index, group in enumerate(seo.find_all('group'), start=1):
                try:
                    self.ws.cell(row=count_row, column=1, value=f'{namber_page}-{index}')
                    self.ws.cell(row=count_row, column=2, value=self.input_request)
                    self.ws.cell(row=count_row, column=4, value='SEO')
                    self.ws.cell(row=count_row, column=5, value=str(group.find('url'))[5:-6])
                    self.ws.cell(row=count_row, column=6, value=self._check_tag(str(group.find('title'))[7:-8]))
                    self.ws.cell(row=count_row, column=7, value=self._check_tag(str(group.find('passage'))[9:-10]))
                    count_row += 1
                    logger.info(f"записано seo - {seo}, по индексу - {namber_page}-{index}")
                except Exception as ex:
                    logger.info(f"Произошла ошибка при записи seo в книгу {ex}")
                    self.handle_parse_error(count_row, 'SEO', namber_page)
                    count_row += 1
        return count_row

    def parse_sitelinks(self, ads, count_row: int):
        """Парсит быстрые ссылки из рекламы и записывает их в рабочую книгу."""
        sitelinks = ads.find_all('sitelink')
        if sitelinks:
            for index, sitelink in enumerate(sitelinks):
                self.ws.cell(row=1, column=8 + index * 2, value=f'Быстрая ссылка - {index + 1}')
                self.ws.cell(row=1, column=9 + index * 2, value=f'Заголовок БС - {index + 1}')
                self.ws.cell(row=count_row, column=8 + index * 2, value=str(sitelink.find('url'))[5:-6])
                self.ws.cell(row=count_row, column=9 + index * 2, value=self._check_tag(str(sitelink.find('title'))[7:-8]))
                logger.info(f"записано sitelink - {sitelink}, по индексу - {index + 1}")

    def handle_parse_error(self, count_row: int, prefix: str, namber_page: str):
        """Обрабатывает ошибки парсинга и записывает их в рабочую книгу."""
        self.ws.cell(row=count_row, column=1, value=f'{prefix}-{namber_page}-1')
        self.ws.cell(row=count_row, column=4, value='В данном разделе информации не найдено')
        self.wb.save(self.name_file)

    @staticmethod
    def _check_tag(value: str) -> str:
        """Проверяет и обрабатывает тег."""
        result = value
        logger.info(f"Обрабатывается тег {result}")
        if "%3Cb%3E" in result:
            result = result.replace("%3Cb%3E", " ").replace("%3C/b%3E", " ")
        if "&amp;nbsp," in result:
            result = result.replace("&amp;nbsp,", " ")
        if "&amp;amp;nbsp," in result:
            result = result.replace("&amp;amp;nbsp,", " ")
        if "&lt;hlword&gt;" in result:
            result = result.replace("&lt;hlword&gt;", " ").replace("&lt;/hlword&gt;", " ")
        if "%3C!-- --%3E" in result:
            result = result.replace("%3C!-- --%3E", " ")
        logger.info(f"Обработан тег {result}")
        return result.strip()

    def add_calculation_sheet(self):
        """Добавляет новый лист с расчетами по поисковым запросам."""
        # Получаем все данные из первого листа
        data = []
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if row[2]:
                logger.info(f"Строки файла ---> {row}")
                data.append(row)

        # Группируем данные по поисковому запросу
        grouped_data = {}
        for row in data:
            request = row[1]  # Поисковый запрос
            site_address = self._extract_main_site(row[4])  # Адрес сайта
            site_weight = row[2]  # Вес сайта

            if request not in grouped_data:
                grouped_data[request] = {}

            if site_address not in grouped_data[request]:
                grouped_data[request][site_address] = 0

            grouped_data[request][site_address] += site_weight

        # Создаем новый лист для каждого поискового запроса
        for request, sites in grouped_data.items():
            new_ws = self.wb.create_sheet(title=request[:30])  # Ограничиваем название листа 30 символами
            new_ws.append(['Адрес сайта', 'Вес сайта', 'Процент от общего веса'])

            total_weight = sum(sites.values())

            for site, weight in sites.items():
                percentage = (weight / total_weight) * 100 if total_weight > 0 else 0
                new_ws.append([site, weight, percentage])

        self.wb.save(self.name_file)
        logger.info("Добавлен новый лист с расчетами по поисковым запросам.")

    def _extract_main_site(self, url: str) -> str:
        """Извлекает основной адрес сайта из полного URL."""

        # Разбираем URL
        parsed_url = urlparse(url)
        domain = parsed_url.netloc

        # Убираем 'www.' если он есть
        if domain.startswith('www.'):
            domain = domain[4:]

        return domain



# Пример использования
# file_gen = FileGenerator(telegram_user_id, advertising_seo, input_requests, input_region)
# file_path = file_gen.generate_file(response)
# file_gen.add_calculation_sheet()
